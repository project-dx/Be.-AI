"""利用者日報・スタッフ日報のExcelダミーデータを取り込むスクリプト。

実行例:
uv run python -m app.seed_from_excel \
  --user-report "C:\\Users\\...\\ダミーアンケート（DX利用者日報データ）.xlsx" \
  --staff-report "C:\\Users\\...\\支援記録（ダミーデータ_3人分）.xlsx"

- 利用者名ごとにユーザーアカウントを作成（既に同名で作成済みならスキップ）
- 担当スタッフは既存の staff@example.com（なければ新規作成）に割り当て
- 取り込み後、スコア再計算とリスク判定を実行する
"""

import argparse
import sys
from datetime import date, datetime, timedelta

import openpyxl
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.database import SessionLocal
from app.core.security import hash_password
from app.models import Profile, StaffDailyReport, User, UserDailyReport
from app.services.risk import evaluate_user_risks
from app.services.scoring import recalculate_range

MOOD_MAP = {
    "無気力": 1,
    "不安": 2,
    "普通": 3,
    "落ち着いている": 4,
    "やる気あり": 5,
}
SLEEP_QUALITY_MAP = {
    "入眠困難": 1,
    "中途覚醒": 2,
    "普通": 3,
    "良い": 5,
}
CONDITION_FATIGUE_MAP = {
    "良好": 1,
    "普通": 2,
    "だるい": 4,
    "眠い": 4,
    "頭痛": 5,
}
BREAKFAST_MAP = {
    "食べた": "eaten",
    "食べていない": "skipped",
}

STAFF_EMAIL = "staff@example.com"
STAFF_PASSWORD = "Staff123!"
STAFF_DISPLAY_NAME = "支援 花子"

IMPORTED_USER_PASSWORD = "User123!"


def _to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def load_user_daily_reports(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, report_date, condition, mood, sleep_quality, breakfast, achievement, activity = row[:8]
        if name is None or report_date is None:
            continue
        rows.append(
            {
                "name": str(name).strip(),
                "report_date": _to_date(report_date),
                "condition": condition,
                "mood": mood,
                "sleep_quality": sleep_quality,
                "breakfast": breakfast,
                "achievement_score": achievement,
                "activity": activity,
            }
        )
    return rows


def load_staff_daily_reports(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, report_date, record_type, content = row[:4]
        if name is None or report_date is None:
            continue
        rows.append(
            {
                "name": str(name).strip(),
                "report_date": _to_date(report_date),
                "record_type": record_type,
                "content": content,
            }
        )
    return rows


def get_or_create_staff(db: Session) -> User:
    staff = db.query(User).filter(User.email == STAFF_EMAIL).first()
    if staff:
        return staff
    staff = User(email=STAFF_EMAIL, password_hash=hash_password(STAFF_PASSWORD), role="staff")
    db.add(staff)
    db.flush()
    db.add(Profile(user_id=staff.id, display_name=STAFF_DISPLAY_NAME))
    db.flush()
    return staff


def get_or_create_user(db: Session, name: str, index: int, staff_id: int) -> User:
    existing = (
        db.query(User)
        .join(Profile, Profile.user_id == User.id)
        .filter(Profile.display_name == name)
        .first()
    )
    if existing:
        return existing
    email = f"imported.user{index}@example.com"
    user = User(email=email, password_hash=hash_password(IMPORTED_USER_PASSWORD), role="user")
    db.add(user)
    db.flush()
    db.add(
        Profile(
            user_id=user.id,
            display_name=name,
            support_start_date=date.today() - timedelta(days=180),
            assigned_staff_id=staff_id,
        )
    )
    db.flush()
    print(f"  作成: {name} -> {email} / {IMPORTED_USER_PASSWORD}")
    return user


def import_user_reports(db: Session, rows: list[dict], name_to_user: dict[str, User]) -> set[int]:
    touched_user_ids: set[int] = set()
    for row in rows:
        user = name_to_user[row["name"]]
        touched_user_ids.add(user.id)
        existing = (
            db.query(UserDailyReport)
            .filter(UserDailyReport.user_id == user.id, UserDailyReport.report_date == row["report_date"])
            .first()
        )
        if existing:
            continue
        achievement_score = row["achievement_score"]
        free_text = f"体調: {row['condition']}" + (
            f" ／ 達成感: {achievement_score}/5" if achievement_score is not None else ""
        )
        db.add(
            UserDailyReport(
                user_id=user.id,
                report_date=row["report_date"],
                mood=MOOD_MAP.get(row["mood"]),
                sleep_quality=SLEEP_QUALITY_MAP.get(row["sleep_quality"]),
                breakfast_status=BREAKFAST_MAP.get(row["breakfast"]),
                fatigue_level=CONDITION_FATIGUE_MAP.get(row["condition"]),
                achievement=row["activity"],
                free_text=free_text,
                is_draft=False,
            )
        )
    return touched_user_ids


def import_staff_reports(db: Session, rows: list[dict], name_to_user: dict[str, User], staff_id: int) -> None:
    for row in rows:
        user = name_to_user[row["name"]]
        is_absence = row["record_type"] == "欠席記録"
        existing = (
            db.query(StaffDailyReport)
            .filter(
                StaffDailyReport.user_id == user.id,
                StaffDailyReport.staff_id == staff_id,
                StaffDailyReport.report_date == row["report_date"],
            )
            .first()
        )
        if existing:
            continue
        db.add(
            StaffDailyReport(
                user_id=user.id,
                staff_id=staff_id,
                report_date=row["report_date"],
                support_content=row["content"],
                urgency="caution" if is_absence else "normal",
            )
        )


def run(user_report_path: str, staff_report_path: str, force: bool = False) -> None:
    settings = get_settings()
    if settings.is_production and not force:
        print("本番環境（ENVIRONMENT=production）では実行しません。--force で強制実行できます")
        sys.exit(1)

    db = SessionLocal()
    try:
        user_rows = load_user_daily_reports(user_report_path)
        staff_rows = load_staff_daily_reports(staff_report_path)

        names = sorted({r["name"] for r in user_rows} | {r["name"] for r in staff_rows})
        print(f"取り込み対象の利用者: {names}")

        staff = get_or_create_staff(db)
        db.flush()

        name_to_user: dict[str, User] = {}
        for i, name in enumerate(names, start=1):
            name_to_user[name] = get_or_create_user(db, name, i, staff.id)
        db.flush()

        touched_user_ids = import_user_reports(db, user_rows, name_to_user)
        import_staff_reports(db, staff_rows, name_to_user, staff.id)
        db.flush()

        all_dates = [r["report_date"] for r in user_rows]
        if all_dates:
            start, end = min(all_dates), max(all_dates)
            for user_id in touched_user_ids:
                recalculate_range(db, user_id, start, end)
            db.flush()
            for user_id in touched_user_ids:
                evaluate_user_risks(db, user_id, end)
            db.flush()

        db.commit()
        print("Excelダミーデータの取り込みが完了しました")
        print("--- 取り込んだ利用者アカウント ---")
        for name, user in name_to_user.items():
            print(f"  {name}: {user.email} / {IMPORTED_USER_PASSWORD}")
        print(f"担当スタッフ: {STAFF_EMAIL} / {STAFF_PASSWORD}")
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excelダミーデータ取り込み")
    parser.add_argument("--user-report", required=True, help="利用者日報Excelファイルのパス")
    parser.add_argument("--staff-report", required=True, help="スタッフ日報（支援記録）Excelファイルのパス")
    parser.add_argument("--force", action="store_true", help="本番環境でも強制実行する")
    args = parser.parse_args()
    run(args.user_report, args.staff_report, force=args.force)
